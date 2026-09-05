
import os
import tracemalloc
import tempfile
from datetime import datetime,timedelta
from collections import namedtuple
from openpyxl import Workbook
from openpyxl.styles import Alignment

from consts import HSP_YEAR_DEFAULT,HSP,HSP_DATES,MEMBERSHIP,MISSING
from utils import load_franks_rac_membership_spreadsheet,appendMsg
from orders import Orders

HSP_ATTENDEE_FIELDS = 'name email expiration_date contribution starbq cabin raffle prv_tent platform order_nums comment'
HspAttendeeTup = namedtuple('HspAttendeeTup', HSP_ATTENDEE_FIELDS)

class HSPAttendees(object):

    def __init__(self,hsp_year=HSP_YEAR_DEFAULT,order_to_debug=None,verbose=False):

        # 9/2/2026. call self.shopifyAndFranksSpreadsheetLoad() to load data.

        self.msg = ''
        self.error = ''
        self.membership_list = []
        self.email_to_rmt_list_map = {}
        self.name_to_rmt_list_map = {}
        self.hat_list = []
        self.contribution_total = 0
        self.starbq_total = 0
        self.cabin_total = 0
        self.raffle_total = 0
        self.prv_tent_total = 0
        self.platform_total = 0
        self.contribution_total_dollars = 0
        self.starbq_total_dollars = 0
        self.cabin_total_dollars = 0
        self.raffle_total_dollars = 0
        self.prv_tent_total_dollars = 0
        self.platform_total_dollars = 0
        self.total_dollars = 0

        self.items_to_share_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'items_to_share')
        if not os.path.isdir(self.items_to_share_dir):
            self.error = f"items_to_share dir of {self.items_to_share_dir}. It must exist and is used to hold Frank's membership Excel files"
            return
        try:
            self.hsp_year = int(hsp_year)
        except:
            self.error = f"hsp_year:'{hsp_year}' passed to HSPAttendee.__init__ is invalid. It must be interpreted as an integer."
            return
        hspDates = HSP_DATES.get(self.hsp_year)
        if not hspDates:
            self.error = f"hsp_year:{self.hsp_year} passed to HSPAttendee.__init__ is invalid. Only supported years are {','.join(str(k) for k in HSP_DATES.keys())}."
            return
        self.created_at_max = (datetime.strptime(hspDates.hsp_end, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')
        self.created_at_min = f'{self.hsp_year}-01-01'
        self.created_at_min_membership = f'{self.hsp_year}-08-01'

        self.hsp_orders = Orders(HSP,verbose=verbose,order_to_debug=order_to_debug)
        if self.hsp_orders.error:
            self.error = self.hsp_orders.error
            return

        # 9/2/2026. XXX You might find it odd that we are loading membership orders when we are supposed to find membership expiration date in Frank's membership files.
        #               The reason we do this is that Frank's files are for end of August. If there are any late membership orders after that but before HSP this will pick them up.
        self.membership_orders = Orders(MEMBERSHIP,verbose=verbose,order_to_debug=order_to_debug)
        if self.membership_orders.error:
            self.error = self.membership_orders.error
            return

        memory_usage = tracemalloc.get_traced_memory()
        print(f'memory usage at exit from HSPAttendee.__init__(...) : current:{memory_usage[0]}, peak:{memory_usage[1]}')

        return

    def buildAndDisplaySales(self):

        if self.error:
            self.error = appendMsg(self.error,'Cannot build and display HSP sales s/s because of preexisting error.')
            return
        if not len(self.hat_list):
            self.error = 'Cannot build and display HSP sales s/s because no HSP sales data loaded.'
            return

            # 9/5/2026. build and display temporary HSP attendee s/s.
        fd, fname = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb = Workbook()
        ws = wb.active
        ws.title = 'HSP Attendees'
        ws.append(HspAttendeeTup._fields)
        ws.freeze_panes = 'A2'

        for hat in self.hat_list:
            ws.append(hat)

        for row in ws.iter_rows():
            for cell in row:
                if cell.value == MISSING:
                    cell.alignment = Alignment(horizontal='right')

        order_nums_col = HspAttendeeTup._fields.index('order_nums') + 1
        for row_num in range(2,ws.max_row + 1):
            cell = ws.cell(row=row_num,column=order_nums_col)
            cell.value = str(cell.value)
            cell.number_format = '@'
            cell.quotePrefix = True

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

        wb.save(fname)
        os.startfile(fname)

        return

    def shopifyAndFranksSpreadsheetLoad(self):

        # 9/2/2026. analogous to the self.ShopifyLoad functions but also loads Frank's membership s/s.

        membership_file_full_path = os.path.join(self.items_to_share_dir,f'RAC AUGUST {self.hsp_year} Membership.xlsx')
        membership_list, email_to_rmt_list_map, name_to_rmt_list_map, self.error, msg = load_franks_rac_membership_spreadsheet(membership_file_full_path)
        self.membership_list.extend(membership_list)
        self.email_to_rmt_list_map.update(email_to_rmt_list_map)
        self.name_to_rmt_list_map.update(name_to_rmt_list_map)
        self.msg = appendMsg(self.msg,msg,print_new_msg=False)
        if self.error:
            return

        self.hsp_orders.shopifyLoad(created_at_min=self.created_at_min,created_at_max=self.created_at_max,order_to_debug=None)
        self.msg = appendMsg(self.msg,self.hsp_orders.msg,print_new_msg=False)
        if self.hsp_orders.error:
            self.error = self.hsp_orders.error
            return

        email_to_owpt_list_map = {}
        name_to_owpt_list_map = {}
        for owpt in self.hsp_orders.full_order_with_properties_dict.values():

            key_email = owpt.email.lower()
            owpt_list = email_to_owpt_list_map.get(key_email,[])
            if not owpt_list:
                email_to_owpt_list_map[key_email] = owpt_list
            owpt_list.append(owpt)

            key_name = owpt.name.lower()
            owpt_list = name_to_owpt_list_map.get(key_name,[])
            if not owpt_list:
                name_to_owpt_list_map[key_name] = owpt_list
            owpt_list.append(owpt)

        self.msg += f"\n# distinct emails in HSP orders:{len(email_to_owpt_list_map)}, # distinct names in HSP orders:{len(name_to_owpt_list_map)}."
        if len(email_to_owpt_list_map) != len(name_to_owpt_list_map):
            self.error = f"MAJOR LOGIC FUCKUP. # distinct emails in HSP orders:{len(email_to_owpt_list_map)} does not equal # distinct names in HSP orders:{len(name_to_owpt_list_map)}"
            return

        # 9/4/2026. populate self.hat_list
        for key_email,owpt_list in email_to_owpt_list_map.items():

            # 9/4/2026. get membership expiration date
            rmt_list = self.email_to_rmt_list_map.get(key_email)
            if rmt_list:
                expiration_date = rmt_list[0].expiration_date
                for rmt in rmt_list:
                    if rmt.expiration_date > expiration_date:
                        expiration_date = rmt.expiration_date
            else:
                expiration_date = MISSING

            name = owpt_list[0].name
            email = owpt_list[0].email
            contribution = 0
            starbq = 0
            cabin = 0
            raffle = 0
            prv_tent = 0
            platform = 0
            contribution_dollars = 0
            starbq_dollars = 0
            cabin_dollars = 0
            raffle_dollars = 0
            prv_tent_dollars = 0
            platform_dollars = 0
            order_nums = ''

            comment = ''
            found_attendee = False
            for owpt in owpt_list:

                onum = owpt.order_num
                toks = onum.split('-')
                if len(toks) == 2:
                    onum = toks[0]
                if not order_nums:
                    order_nums = onum
                else:
                    if onum not in order_nums.split('|'):
                        order_nums += '|' + onum

                if owpt.name not in name.split('|'):
                    name += '|' + owpt.name
                name = owpt.name

                if owpt.sku == 'hsp_contribution':
                    found_attendee = True
                    contribution += owpt.quantity
                    contribution_dollars += int(owpt.paid)
                if owpt.sku == 'hsp_starbq':
                    found_attendee = True
                    starbq += owpt.quantity
                    starbq_dollars += int(owpt.paid)
                if owpt.sku.startswith('hsp_cabin_for_2_'):
                    found_attendee = True
                    cabin += owpt.quantity
                    cabin_dollars += int(owpt.paid)
                if owpt.sku.startswith('hsp_raffle_'):
                    raffle += owpt.quantity * int(owpt.sku.split('_')[-1])
                    raffle_dollars += int(owpt.paid)
                if owpt.sku.startswith('hsp_prv_tent_'):
                    found_attendee = True
                    prv_tent += owpt.quantity
                    prv_tent_dollars += int(owpt.paid)
                if owpt.sku.startswith('hsp_platform_for_'):
                    found_attendee = True
                    platform += owpt.quantity
                    platform_dollars += int(owpt.paid)

            if not found_attendee:
                contribution = MISSING
                starbq = MISSING

            if isinstance(contribution,int):
                self.contribution_total += contribution
                self.contribution_total_dollars += contribution_dollars
            if isinstance(starbq,int):
                self.starbq_total += starbq
                self.starbq_total_dollars += starbq_dollars
            self.raffle_total += raffle
            self.raffle_total_dollars += raffle_dollars
            self.cabin_total += cabin
            self.cabin_total_dollars += cabin_dollars
            self.prv_tent_total += prv_tent
            self.prv_tent_total_dollars += prv_tent_dollars
            self.platform_total += platform
            self.platform_total_dollars += platform_dollars
            self.hat_list.append(HspAttendeeTup(name,email,expiration_date,contribution,starbq,cabin,raffle,prv_tent,platform,order_nums,comment))

        self.total_dollars = self.contribution_total_dollars + self.starbq_total_dollars + self.raffle_total_dollars + self.cabin_total_dollars + self.prv_tent_total_dollars + \
                             self.platform_total_dollars

        self.membership_orders.shopifyLoad(created_at_min=self.created_at_min_membership,created_at_max=self.created_at_max,order_to_debug=None)
        if self.membership_orders.error:
            self.error = self.membership_orders.error
            return

        self.msg = appendMsg(self.msg,self.membership_orders.msg,print_new_msg=False)
        dt_str = datetime.now().strftime('%Y-%m-%d %#I:%M%p')
        msg_totals = f"HSP SALES TOTALS as of {dt_str}:\ncontribution:{self.contribution_total}, starbq:{self.starbq_total}, raffle:{self.raffle_total}, cabin:{self.cabin_total}, " \
                     f"prv_tent:{self.prv_tent_total}, platform:{self.platform_total}"
        msg_totals += f"\nHSP SALES TOTALS DOLLARS\ncontribution:${self.contribution_total_dollars}, starbq:${self.starbq_total_dollars}, raffle:${self.raffle_total_dollars}," \
                     f" cabin:${self.cabin_total_dollars}, prv_tent:${self.prv_tent_total_dollars}, platform:${self.platform_total_dollars}"
        msg_totals += f"\nTOTAL SALES: ${self.total_dollars}"
        print(msg_totals)
        self.msg = appendMsg(self.msg,msg_totals,print_new_msg=False)

        return